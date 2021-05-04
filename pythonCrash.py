import openpyxl
from openpyxl.chart import (PieChart, Reference, Series)

#Load an existing workbook to manipulate
workbook_object = openpyxl.load_workbook('crashData.xlsx')

#Rename the first sheet as 'Originale data'
worksheet_object0 = workbook_object.active
worksheet_object0.title = 'Original Data'

#Create a new worksheet object and call it 'Filtered data'
worksheet_object1 = workbook_object.create_sheet('Filtered Data')

#Copy 'Original Data' to 'Filtered Data' worksheet
for i in range(1, worksheet_object0.max_row + 1):
    for j in range(1, worksheet_object0.max_column + 1):
        worksheet_object1.cell(row=i, column=j).value = worksheet_object0.cell(row=i, column=j).value

#Remove all the unwanted columns from 'Filtered Data' worksheet
unwanted_columns = ['AccidentNumber', 'County', 'RouteType', 'Route', 'Milelog', 'IntersectingRoute',
                  'RampSection', 'DistanceFrom', 'DirectionFrom', 'LocationOfImpact', 'FirstHarmfulEvent',
                  'DirVeh1', 'DirVeh2', 'MnvrVeh1', 'MnvrVeh2', 'MicrofilmNo', 'U1Factors',	'U2Factors',
                  'Vendor', 'IntersectRouteType', 'U1FirstHarmfulEvent', 'U2FirstHarmfulEvent',
                  ]

#delet unwated columns
for i in range(1, worksheet_object1.max_column + 1):
    for column in unwanted_columns:
        if worksheet_object1.cell(row=1, column=i).value == column:
            worksheet_object1.delete_cols(i)

#Store crash data into a dictionary
crash_data = {'Angle': 0,
              'Side-Swipe': 0,
              'Rear End': 0,
              'Head On': 0,
              }

##Populate dictionary with data from worksheet
for i in range(2, worksheet_object1.max_row + 1):
    manner_of_collision = worksheet_object1.cell(row=i, column=6).value
    if manner_of_collision in crash_data:
        crash_data[manner_of_collision] += 1

#Convert data into a list
data = [['Manner of Collision', 'Number of Crash']]
for key, value in crash_data.items():
    data.append([key, value])

#Create new worksheet object for the pie chart
worksheet_object2 = workbook_object.create_sheet('Pie Chart')
for row in data:
    worksheet_object2.append(row)

#Create reference object
reference_object = Reference(worksheet_object2, min_col=2, min_row=2, max_col=2, max_row=5)

#Create series object
labels = Reference(worksheet_object2, min_col=1, min_row=2, max_col=1, max_row=5)
series_object = Series(reference_object)

#Create chart object
chart_object = PieChart()

#Append series object to chart object
chart_object.append(series_object)
chart_object.set_categories(labels)
chart_object.title = worksheet_object2['A1'].value
#Append chart object to worksheet
worksheet_object2.add_chart(chart_object)


workbook_object.save('crashData.xlsx')